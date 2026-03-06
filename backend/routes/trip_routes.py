"""
Trip routes for Servex Holdings backend.
Handles trip CRUD operations, trip details, and expense management.
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import io

from database import db
from dependencies import get_current_user, get_tenant_id, build_warehouse_filter
from models.schemas import Trip, TripCreate, TripUpdate, TripExpense, TripExpenseCreate, TripExpenseUpdate, create_audit_log
from models.enums import TripStatus, ExpenseCategory, AuditAction
from services.trip_service import generate_trip_prefix, get_warehouse_code
from services.barcode_service import generate_barcode

router = APIRouter()

@router.get("/trips/next-number")
async def get_next_trip_number(tenant_id: str = Depends(get_tenant_id)):
    """Get the next sequential trip number for creating a new trip"""
    # Find the latest trip number starting with 'S' followed by 1-4 digits only
    trips = await db.trips.find(
        {"tenant_id": tenant_id, "trip_number": {"$regex": "^S\\d{1,4}$"}},
        {"trip_number": 1, "_id": 0}
    ).to_list(1000)
    
    max_num = 0
    for trip in trips:
        try:
            num = int(trip["trip_number"][1:])
            if num > max_num:
                max_num = num
        except (ValueError, IndexError):
            continue
    
    next_number = f"S{max_num + 1}"
    return {"next_trip_number": next_number}

@router.get("/trips/next-number-by-warehouse")
async def get_next_trip_number_by_warehouse(
    warehouse_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Preview next trip number for a given warehouse without incrementing counter"""
    warehouse = await db.warehouses.find_one({"id": warehouse_id, "tenant_id": tenant_id}, {"_id": 0, "name": 1})
    if not warehouse:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    warehouse_name = warehouse["name"]
    warehouse_code = get_warehouse_code(warehouse_name)
    year = datetime.now(timezone.utc).strftime("%y")
    counter_key = f"trip_seq_{tenant_id}_{warehouse_code}_{year}"
    counter_doc = await db.counters.find_one({"key": counter_key})
    current_val = counter_doc["value"] if counter_doc else 0
    next_seq = current_val + 1
    next_prefix = f"{warehouse_code}-{str(next_seq).zfill(2)}-{year}"
    return {"next_trip_number": next_prefix, "warehouse_code": warehouse_code, "warehouse_name": warehouse_name}


@router.get("/trips")
async def list_trips(
    status: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """List all trips. Status can be comma-separated for multiple values (e.g., 'planning,loading').
    
    SECURITY: For users with warehouse restrictions, only show trips that have parcels
    from their allowed warehouses.
    """
    query = {"tenant_id": tenant_id}
    if status:
        # Support comma-separated status values
        if "," in status:
            query["status"] = {"$in": status.split(",")}
        else:
            query["status"] = status
    
    # SECURITY: Apply warehouse-based filtering
    warehouse_filter = build_warehouse_filter(user)
    if warehouse_filter:
        # Find trip IDs that have parcels from user's allowed warehouses
        allowed_warehouses = user.get("allowed_warehouses", [])
        trip_ids_cursor = await db.shipments.distinct(
            "trip_id",
            {"tenant_id": tenant_id, "warehouse_id": {"$in": allowed_warehouses}}
        )
        # Also include trips with destination_warehouse in allowed list
        dest_trip_ids = await db.trips.distinct(
            "id",
            {"tenant_id": tenant_id, "destination_warehouse_id": {"$in": allowed_warehouses}}
        )
        all_allowed_trips = list(set(trip_ids_cursor + dest_trip_ids))
        query["id"] = {"$in": all_allowed_trips}
    
    trips = await db.trips.find(query, {"_id": 0}).sort("departure_date", -1).to_list(100)
    return trips

@router.get("/trips/{trip_id}")
async def get_trip(trip_id: str, tenant_id: str = Depends(get_tenant_id)):
    """Get single trip with shipments and expenses"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(100)
    
    expenses = await db.trip_expenses.find(
        {"trip_id": trip_id},
        {"_id": 0}
    ).sort("expense_date", -1).to_list(100)
    
    # Calculate expense totals by category
    expense_totals = {}
    total_expenses = 0
    for expense in expenses:
        category = expense.get("category", "other")
        amount = expense.get("amount", 0)
        expense_totals[category] = expense_totals.get(category, 0) + amount
        total_expenses += amount
    
    return {
        **trip,
        "shipments": shipments,
        "expenses": expenses,
        "expense_totals": expense_totals,
        "total_expenses": total_expenses
    }

@router.post("/trips", response_model=Trip)
async def create_trip(
    request: Request,
    trip_data: TripCreate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Create a new trip with trip-based numbering (SESSION N Part 1)"""
    trip_dict = trip_data.model_dump()

    # Determine warehouse name for trip number generation
    departure_warehouse_id = trip_dict.pop("departure_warehouse_id", None)
    warehouse_name = None
    if departure_warehouse_id:
        wh = await db.warehouses.find_one({"id": departure_warehouse_id, "tenant_id": tenant_id}, {"_id": 0, "name": 1})
        if wh:
            warehouse_name = wh["name"]

    # Fallback to route[0] then trip_number
    route = trip_dict.get("route", [])
    if not warehouse_name:
        warehouse_name = route[0] if route else (trip_dict.get("trip_number") or "Warehouse")

    # Always auto-generate trip prefix (ignore any manually entered trip_number)
    prefix_data = await generate_trip_prefix(tenant_id, warehouse_name)
    trip_dict["trip_number"] = prefix_data["trip_prefix"]
    trip_dict["trip_prefix"] = prefix_data["trip_prefix"]
    trip_dict["warehouse_code"] = prefix_data["warehouse_code"]
    trip_dict["trip_sequence"] = prefix_data["trip_sequence"]
    trip_dict["year"] = prefix_data["year"]
    trip_dict["invoice_seq"] = 0

    # Check trip_number uniqueness within tenant
    existing = await db.trips.find_one({
        "tenant_id": tenant_id,
        "trip_number": trip_dict["trip_number"]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Trip number already exists for this tenant")

    trip = Trip(
        **trip_dict,
        tenant_id=tenant_id,
        created_by=user["id"]
    )

    doc = trip.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('locked_at'):
        doc['locked_at'] = doc['locked_at'].isoformat()
    # Persist extra fields
    for key in ["trip_prefix", "warehouse_code", "trip_sequence", "year", "invoice_seq"]:
        if key in trip_dict:
            doc[key] = trip_dict[key]
    await db.trips.insert_one(doc)

    # Audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.create,
        table_name="trips",
        record_id=trip.id,
        new_value=doc,
        ip_address=request.client.host if request.client else None
    )

    return trip

@router.put("/trips/{trip_id}")
async def update_trip(
    request: Request,
    trip_id: str,
    update_data: TripUpdate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Update trip - handles locking when status changes to 'closed'"""
    # Get current trip
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    old_trip = dict(trip)
    
    # Check if trip is locked (only owner can modify locked trips)
    if trip.get("locked_at") and user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Trip is locked. Only owner can modify.")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    # Check if trip_number is being changed and ensure uniqueness
    if "trip_number" in update_dict and update_dict["trip_number"] != trip.get("trip_number"):
        existing = await db.trips.find_one({
            "tenant_id": tenant_id,
            "trip_number": update_dict["trip_number"],
            "id": {"$ne": trip_id}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Trip number already exists for this tenant")
    
    # Determine action type
    action = AuditAction.status_change if "status" in update_dict else AuditAction.update
    
    # Handle status change to 'closed' - set locked_at timestamp
    if update_dict.get("status") == "closed" and trip.get("status") != "closed":
        update_dict["locked_at"] = datetime.now(timezone.utc).isoformat()
    
    # Handle status change to 'in_transit' - set actual_departure timestamp
    # Also unassign parcels that were assigned but never loaded (still in 'staged' status)
    if update_dict.get("status") == "in_transit" and trip.get("status") != "in_transit":
        if not trip.get("actual_departure"):
            update_dict["actual_departure"] = datetime.now(timezone.utc).isoformat()
        
        # S4: Bulk-update all 'loaded' parcels for this trip to 'in_transit'
        await db.shipments.update_many(
            {
                "trip_id": trip_id,
                "tenant_id": tenant_id,
                "status": "loaded"
            },
            {
                "$set": {"status": "in_transit"}
            }
        )
        
        # Unassign parcels that are still in 'staged' status (not loaded onto the truck)
        # These parcels were assigned to the trip but never actually scanned/loaded
        await db.shipments.update_many(
            {
                "trip_id": trip_id,
                "tenant_id": tenant_id,
                "status": "staged"  # Only parcels that were never loaded
            },
            {
                "$set": {
                    "trip_id": None,
                    "status": "warehouse"  # Return to warehouse status
                }
            }
        )
    
    # Handle status change to 'delivered' - set actual_arrival timestamp
    if update_dict.get("status") == "delivered" and trip.get("status") != "delivered":
        if not trip.get("actual_arrival"):
            update_dict["actual_arrival"] = datetime.now(timezone.utc).isoformat()
    
    if update_dict:
        await db.trips.update_one(
            {"id": trip_id, "tenant_id": tenant_id},
            {"$set": update_dict}
        )
    
    new_trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    
    # Audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=action,
        table_name="trips",
        record_id=trip_id,
        old_value=old_trip,
        new_value=new_trip,
        ip_address=request.client.host if request.client else None
    )
    
    return new_trip

@router.delete("/trips/{trip_id}")
async def delete_trip(
    request: Request,
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Delete a trip (only if not locked or user is owner)"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    old_trip = dict(trip)
    
    if trip.get("locked_at") and user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Cannot delete locked trip")
    
    # Unassign all shipments from this trip
    await db.shipments.update_many(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"$set": {"trip_id": None, "status": "warehouse"}}
    )
    
    # Delete associated expenses
    await db.trip_expenses.delete_many({"trip_id": trip_id})
    
    # Delete trip
    await db.trips.delete_one({"id": trip_id, "tenant_id": tenant_id})
    
    # Audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.delete,
        table_name="trips",
        record_id=trip_id,
        old_value=old_trip,
        ip_address=request.client.host if request.client else None
    )
    
    return {"message": "Trip deleted"}

@router.post("/trips/{trip_id}/assign-shipment/{shipment_id}")
async def assign_shipment_to_trip(
    trip_id: str,
    shipment_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Assign a shipment to a trip"""
    # Verify trip exists and is not locked
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("locked_at"):
        raise HTTPException(status_code=403, detail="Cannot modify shipments on a locked trip")
    
    # Determine destination from trip route
    trip_destination = None
    if trip.get("route") and len(trip["route"]) > 0:
        trip_destination = trip["route"][-1]

    # Update shipment with destination
    shipment_update = {"trip_id": trip_id, "status": "staged"}
    if trip_destination:
        shipment_update["destination"] = trip_destination

    result = await db.shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id},
        {"$set": shipment_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    # Update piece barcodes with new trip number
    pieces = await db.shipment_pieces.find({"shipment_id": shipment_id}, {"_id": 0}).to_list(100)
    
    shipment_count = await db.shipments.count_documents({
        "tenant_id": tenant_id,
        "trip_id": trip_id
    })
    
    for piece in pieces:
        new_barcode = generate_barcode(trip["trip_number"], shipment_count, piece["piece_number"])
        await db.shipment_pieces.update_one(
            {"id": piece["id"]},
            {"$set": {"barcode": new_barcode}}
        )
    
    return {"message": "Shipment assigned to trip"}

@router.post("/trips/{trip_id}/unassign-shipment/{shipment_id}")
async def unassign_shipment_from_trip(
    trip_id: str,
    shipment_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Remove a shipment from a trip"""
    # Verify trip exists and is not locked
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("locked_at"):
        raise HTTPException(status_code=403, detail="Cannot modify shipments on a locked trip")
    
    # Update shipment
    result = await db.shipments.update_one(
        {"id": shipment_id, "tenant_id": tenant_id, "trip_id": trip_id},
        {"$set": {"trip_id": None, "status": "warehouse"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Shipment not found or not assigned to this trip")
    
    # Update piece barcodes back to TEMP format
    pieces = await db.shipment_pieces.find({"shipment_id": shipment_id}, {"_id": 0}).to_list(100)
    
    for piece in pieces:
        new_barcode = generate_barcode(None, 0, piece["piece_number"])
        await db.shipment_pieces.update_one(
            {"id": piece["id"]},
            {"$set": {"barcode": new_barcode}}
        )
    
    return {"message": "Shipment removed from trip"}

@router.get("/trips/{trip_id}/summary")
async def get_trip_summary(
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get comprehensive trip summary with statistics"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Get vehicle details if assigned
    vehicle = None
    if trip.get("vehicle_id"):
        vehicle = await db.vehicles.find_one({"id": trip["vehicle_id"]}, {"_id": 0})
    
    # Get driver details if assigned
    driver = None
    if trip.get("driver_id"):
        driver = await db.drivers.find_one({"id": trip["driver_id"]}, {"_id": 0})
    
    # Get shipments assigned to this trip
    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Calculate stats
    total_parcels = len(shipments)
    total_weight = sum(s.get("total_weight", 0) or 0 for s in shipments)
    unique_clients = set(s.get("client_id") for s in shipments if s.get("client_id"))
    
    # Count pieces
    total_pieces = 0
    for shipment in shipments:
        pieces = await db.shipment_pieces.count_documents({"shipment_id": shipment["id"]})
        total_pieces += pieces
    
    # Count loaded parcels (status in ['staged', 'loaded', 'in_transit', 'delivered'])
    loaded_statuses = ['staged', 'loaded', 'in_transit', 'delivered']
    loaded_parcels = sum(1 for s in shipments if s.get("status") in loaded_statuses)
    loading_percentage = round((loaded_parcels / total_parcels * 100) if total_parcels > 0 else 0)
    
    # Get invoiced value - query invoices by BOTH trip_id and shipment_ids
    shipment_ids = [s["id"] for s in shipments]
    
    # First get invoices linked directly by trip_id
    trip_invoices = await db.invoices.find(
        {"tenant_id": tenant_id, "trip_id": trip_id},
        {"id": 1, "total": 1, "_id": 0}
    ).to_list(1000)
    
    # Also get invoices linked by shipment_ids (for backward compatibility)
    shipment_invoices = await db.invoices.find(
        {"tenant_id": tenant_id, "shipment_ids": {"$in": shipment_ids}},
        {"id": 1, "total": 1, "_id": 0}
    ).to_list(1000) if shipment_ids else []
    
    # Combine and deduplicate
    seen_invoice_ids = set()
    invoiced_value = 0
    for inv in trip_invoices + shipment_invoices:
        if inv["id"] not in seen_invoice_ids:
            seen_invoice_ids.add(inv["id"])
            invoiced_value += inv.get("total", 0) or 0
    
    # Get created by user
    created_by_user = None
    if trip.get("created_by"):
        created_by_user = await db.users.find_one({"id": trip["created_by"]}, {"name": 1, "_id": 0})
    
    return {
        "trip": {
            **trip,
            "vehicle": vehicle,
            "driver": driver
        },
        "stats": {
            "total_parcels": total_parcels,
            "total_pieces": total_pieces,
            "total_weight": round(total_weight, 2),
            "total_clients": len(unique_clients),
            "invoiced_value": round(invoiced_value, 2),
            "loaded_parcels": loaded_parcels,
            "loading_percentage": loading_percentage
        },
        "created_by": created_by_user.get("name") if created_by_user else None,
        "created_at": trip.get("created_at")
    }

@router.get("/trips-with-stats")
async def list_trips_with_stats(
    status: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id)
):
    """List all trips with calculated statistics"""
    query = {"tenant_id": tenant_id}
    if status and status != "all":
        query["status"] = status
    
    trips = await db.trips.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    result = []
    for trip in trips:
        # Get shipments for this trip
        shipments = await db.shipments.find(
            {"trip_id": trip["id"], "tenant_id": tenant_id},
            {"_id": 0, "id": 1, "total_weight": 1, "client_id": 1, "status": 1}
        ).to_list(2000)
        
        total_parcels = len(shipments)
        total_weight = sum(s.get("total_weight", 0) or 0 for s in shipments)
        unique_clients = set(s.get("client_id") for s in shipments if s.get("client_id"))
        
        # Count loaded
        loaded_statuses = ['staged', 'loaded', 'in_transit', 'delivered']
        loaded_parcels = sum(1 for s in shipments if s.get("status") in loaded_statuses)
        loading_percentage = round((loaded_parcels / total_parcels * 100) if total_parcels > 0 else 0)
        
        # Get invoiced value - query by BOTH trip_id and shipment_ids
        shipment_ids = [s["id"] for s in shipments]
        
        # Get invoices linked by trip_id
        trip_invoices = await db.invoices.find(
            {"tenant_id": tenant_id, "trip_id": trip["id"]},
            {"id": 1, "total": 1, "_id": 0}
        ).to_list(1000)
        
        # Also get invoices linked by shipment_ids
        shipment_invoices = []
        if shipment_ids:
            shipment_invoices = await db.invoices.find(
                {"tenant_id": tenant_id, "shipment_ids": {"$in": shipment_ids}},
                {"id": 1, "total": 1, "_id": 0}
            ).to_list(1000)
        
        # Combine and deduplicate
        seen_invoice_ids = set()
        invoiced_value = 0
        for inv in trip_invoices + shipment_invoices:
            if inv["id"] not in seen_invoice_ids:
                seen_invoice_ids.add(inv["id"])
                invoiced_value += inv.get("total", 0) or 0
        
        # Get vehicle and driver info
        vehicle = None
        vehicle_capacity_kg = trip.get("capacity_kg") or 0
        vehicle_capacity_cbm = trip.get("capacity_cbm") or 0
        if trip.get("vehicle_id"):
            vehicle = await db.vehicles.find_one({"id": trip["vehicle_id"]}, {"_id": 0, "registration_number": 1, "vehicle_type": 1, "max_weight_kg": 1, "max_volume_cbm": 1, "name": 1})
            # Use vehicle capacity as fallback if trip doesn't have it set
            if vehicle:
                if not vehicle_capacity_kg and vehicle.get("max_weight_kg"):
                    vehicle_capacity_kg = vehicle["max_weight_kg"]
                if not vehicle_capacity_cbm and vehicle.get("max_volume_cbm"):
                    vehicle_capacity_cbm = vehicle["max_volume_cbm"]
        
        driver = None
        if trip.get("driver_id"):
            driver = await db.drivers.find_one({"id": trip["driver_id"]}, {"_id": 0, "name": 1, "phone": 1})
        
        # Calculate total CBM for parcels
        cbm_shipments = await db.shipments.find(
            {"trip_id": trip["id"], "tenant_id": tenant_id},
            {"_id": 0, "total_cbm": 1}
        ).to_list(2000)
        total_cbm = sum(s.get("total_cbm", 0) or 0 for s in cbm_shipments)
        
        result.append({
            **trip,
            "vehicle": vehicle,
            "driver": driver,
            "stats": {
                "total_parcels": total_parcels,
                "total_weight": round(total_weight, 2),
                "total_cbm": round(total_cbm, 4),
                "total_clients": len(unique_clients),
                "invoiced_value": round(invoiced_value, 2),
                "loaded_parcels": loaded_parcels,
                "loading_percentage": loading_percentage,
                "capacity_kg": vehicle_capacity_kg,
                "capacity_cbm": vehicle_capacity_cbm
            }
        })
    
    return result

# ============ TRIP EXPENSES ROUTES ============

@router.get("/trips/{trip_id}/expenses", response_model=List[TripExpense])
async def list_trip_expenses(
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """List all expenses for a trip"""
    # Verify trip exists and belongs to tenant
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    expenses = await db.trip_expenses.find(
        {"trip_id": trip_id},
        {"_id": 0}
    ).sort("expense_date", -1).to_list(100)
    
    return expenses

@router.post("/trips/{trip_id}/expenses", response_model=TripExpense)
async def create_trip_expense(
    trip_id: str,
    expense_data: TripExpenseCreate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Add an expense to a trip"""
    # Verify trip exists and belongs to tenant
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    expense = TripExpense(
        **expense_data.model_dump(),
        trip_id=trip_id,
        created_by=user["id"]
    )
    
    doc = expense.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.trip_expenses.insert_one(doc)
    
    return expense

@router.put("/trips/{trip_id}/expenses/{expense_id}")
async def update_trip_expense(
    trip_id: str,
    expense_id: str,
    update_data: TripExpenseUpdate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Update a trip expense (locked trips: owner only)"""
    # Verify trip exists and belongs to tenant
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check if trip is locked (only owner can edit expenses on locked trips)
    if trip.get("locked_at") and user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Trip is locked. Only owner can edit expenses.")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if update_dict:
        result = await db.trip_expenses.update_one(
            {"id": expense_id, "trip_id": trip_id},
            {"$set": update_dict}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Expense not found")
    
    expense = await db.trip_expenses.find_one({"id": expense_id, "trip_id": trip_id}, {"_id": 0})
    return expense

@router.delete("/trips/{trip_id}/expenses/{expense_id}")
async def delete_trip_expense(
    trip_id: str,
    expense_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Delete a trip expense (locked trips: owner only)"""
    # Verify trip exists and belongs to tenant
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check if trip is locked
    if trip.get("locked_at") and user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Trip is locked. Only owner can delete expenses.")
    
    result = await db.trip_expenses.delete_one({"id": expense_id, "trip_id": trip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Expense deleted"}

@router.get("/trips/{trip_id}/parcels")
async def get_trip_parcels(
    trip_id: str,
    status: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get all parcels assigned to a trip with detailed info"""
    # Verify trip exists
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Build query
    query = {"trip_id": trip_id, "tenant_id": tenant_id}
    if status and status != "all":
        if status == "not_loaded":
            query["status"] = {"$in": ["warehouse", "staged"]}
        elif status == "loaded":
            query["status"] = "loaded"
        elif status == "delivered":
            query["status"] = "delivered"
    
    parcels = await db.shipments.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    
    # Enrich with client names and piece counts
    result = []
    for parcel in parcels:
        client = await db.clients.find_one({"id": parcel.get("client_id")}, {"_id": 0, "name": 1})
        pieces = await db.shipment_pieces.find({"shipment_id": parcel["id"]}, {"_id": 0}).to_list(100)
        
        result.append({
            **parcel,
            "client_name": client.get("name") if client else "Unknown",
            "pieces": pieces,
            "piece_count": len(pieces)
        })
    
    return result

@router.get("/trips/{trip_id}/clients-summary")
async def get_trip_clients_summary(
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get client summary for a trip with invoice info"""
    # Verify trip exists
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Get all parcels for this trip
    parcels = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(2000)
    
    # Group by client
    client_data = {}
    for parcel in parcels:
        client_id = parcel.get("client_id")
        if not client_id:
            continue
        
        if client_id not in client_data:
            client = await db.clients.find_one({"id": client_id}, {"_id": 0})
            client_data[client_id] = {
                "client_id": client_id,
                "client_name": client.get("name") if client else "Unknown",
                "client_phone": client.get("phone") if client else None,
                "parcel_count": 0,
                "total_weight": 0,
                "shipment_ids": [],
                "invoices": []
            }
        
        client_data[client_id]["parcel_count"] += 1
        client_data[client_id]["total_weight"] += parcel.get("total_weight", 0) or 0
        client_data[client_id]["shipment_ids"].append(parcel["id"])
    
    # Get invoices for this trip (query by BOTH trip_id and shipment_ids)
    trip_invoices = await db.invoices.find(
        {"tenant_id": tenant_id, "trip_id": trip_id},
        {"_id": 0}
    ).to_list(100)
    
    # Build a map of client_id -> invoices from trip_invoices
    trip_invoice_map = {}
    for inv in trip_invoices:
        cid = inv.get("client_id")
        if cid:
            if cid not in trip_invoice_map:
                trip_invoice_map[cid] = []
            trip_invoice_map[cid].append(inv)
    
    # For each client, add their invoices (from trip_id query OR shipment_ids query)
    for client_id, data in client_data.items():
        # First check invoices linked by trip_id
        client_invoices = trip_invoice_map.get(client_id, [])
        
        # Also check invoices linked by shipment_ids (for backward compatibility)
        if data["shipment_ids"]:
            shipment_invoices = await db.invoices.find(
                {"tenant_id": tenant_id, "client_id": client_id, "shipment_ids": {"$in": data["shipment_ids"]}},
                {"_id": 0}
            ).to_list(100)
            # Add any invoices not already in the list
            existing_ids = {inv["id"] for inv in client_invoices}
            for inv in shipment_invoices:
                if inv["id"] not in existing_ids:
                    client_invoices.append(inv)
        
        for inv in client_invoices:
            payments = await db.payments.find({"invoice_id": inv["id"]}, {"_id": 0, "amount": 1}).to_list(100)
            paid_amount = sum(p.get("amount", 0) for p in payments)
            data["invoices"].append({
                "id": inv["id"],
                "invoice_number": inv.get("invoice_number"),
                "total": inv.get("total", 0),
                "status": inv.get("status"),
                "paid_amount": paid_amount
            })
    
    # Also add any clients from trip invoices that aren't in the parcels list
    for client_id, invs in trip_invoice_map.items():
        if client_id not in client_data:
            client = await db.clients.find_one({"id": client_id}, {"_id": 0})
            client_data[client_id] = {
                "client_id": client_id,
                "client_name": client.get("name") if client else "Unknown",
                "client_phone": client.get("phone") if client else None,
                "parcel_count": 0,
                "total_weight": 0,
                "shipment_ids": [],
                "invoices": []
            }
            for inv in invs:
                payments = await db.payments.find({"invoice_id": inv["id"]}, {"_id": 0, "amount": 1}).to_list(100)
                paid_amount = sum(p.get("amount", 0) for p in payments)
                client_data[client_id]["invoices"].append({
                    "id": inv["id"],
                    "invoice_number": inv.get("invoice_number"),
                    "total": inv.get("total", 0),
                    "status": inv.get("status"),
                    "paid_amount": paid_amount
                })
    
    # Calculate totals
    result = list(client_data.values())
    totals = {
        "total_clients": len(result),
        "total_parcels": sum(c["parcel_count"] for c in result),
        "total_weight": round(sum(c["total_weight"] for c in result), 2),
        "total_invoiced": sum(sum(inv["total"] for inv in c["invoices"]) for c in result),
        "total_paid": sum(sum(inv["paid_amount"] for inv in c["invoices"]) for c in result)
    }
    
    return {"clients": result, "totals": totals}

@router.get("/trips/{trip_id}/history")
async def get_trip_history(
    trip_id: str,
    filter_type: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get audit history for a trip and related records"""
    # Verify trip exists
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Get shipment IDs for this trip
    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"id": 1, "_id": 0}
    ).to_list(2000)
    shipment_ids = [s["id"] for s in shipments]
    
    # Build query for audit logs
    queries = [{"record_id": trip_id, "table_name": "trips"}]
    
    if filter_type == "parcels" or filter_type is None:
        for sid in shipment_ids:
            queries.append({"record_id": sid, "table_name": "shipments"})
    
    if filter_type == "expenses" or filter_type is None:
        expenses = await db.trip_expenses.find({"trip_id": trip_id}, {"id": 1, "_id": 0}).to_list(100)
        for exp in expenses:
            queries.append({"record_id": exp["id"], "table_name": "trip_expenses"})
    
    if filter_type == "invoices" or filter_type is None:
        invoices = await db.invoices.find(
            {"tenant_id": tenant_id, "shipment_ids": {"$in": shipment_ids}},
            {"id": 1, "_id": 0}
        ).to_list(100)
        for inv in invoices:
            queries.append({"record_id": inv["id"], "table_name": "invoices"})
    
    # Get audit logs
    audit_logs = await db.audit_logs.find(
        {"$or": queries, "tenant_id": tenant_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    
    # Enrich with user names
    result = []
    user_cache = {}
    for log in audit_logs:
        user_id = log.get("user_id")
        if user_id and user_id not in user_cache:
            user = await db.users.find_one({"id": user_id}, {"name": 1, "_id": 0})
            user_cache[user_id] = user.get("name") if user else "Unknown"
        
        result.append({
            **log,
            "user_name": user_cache.get(user_id, "System")
        })
    
    return result

@router.post("/trips/{trip_id}/close")
async def close_trip(
    trip_id: str,
    request: Request,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Close/lock a trip (owner only)"""
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Only owner can close trips")
    
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("locked_at"):
        raise HTTPException(status_code=400, detail="Trip is already closed")
    
    from datetime import datetime, timezone
    locked_at = datetime.now(timezone.utc).isoformat()
    
    old_value = dict(trip)
    await db.trips.update_one(
        {"id": trip_id},
        {"$set": {"status": "closed", "locked_at": locked_at}}
    )
    
    # Audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.update,
        table_name="trips",
        record_id=trip_id,
        old_value=old_value,
        new_value={"status": "closed", "locked_at": locked_at},
        ip_address=request.client.host if request.client else None
    )
    
    return {"message": f"Trip {trip.get('trip_number')} closed successfully", "locked_at": locked_at}

@router.delete("/trips/{trip_id}/parcels/{parcel_id}")
async def remove_parcel_from_trip(
    trip_id: str,
    parcel_id: str,
    request: Request,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Remove a parcel from a trip (unassign)"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("locked_at"):
        raise HTTPException(status_code=403, detail="Cannot modify closed trip")
    
    parcel = await db.shipments.find_one({"id": parcel_id, "trip_id": trip_id}, {"_id": 0})
    if not parcel:
        raise HTTPException(status_code=404, detail="Parcel not found on this trip")
    
    old_value = dict(parcel)
    await db.shipments.update_one(
        {"id": parcel_id},
        {"$set": {"trip_id": None, "status": "warehouse"}}
    )
    
    # Reset barcodes to TEMP
    pieces = await db.shipment_pieces.find({"shipment_id": parcel_id}, {"_id": 0}).to_list(100)
    for piece in pieces:
        temp_barcode = f"TEMP-{uuid.uuid4().hex[:8].upper()}"
        await db.shipment_pieces.update_one(
            {"id": piece["id"]},
            {"$set": {"barcode": temp_barcode}}
        )
    
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.update,
        table_name="shipments",
        record_id=parcel_id,
        old_value=old_value,
        new_value={"trip_id": None, "status": "warehouse"},
        ip_address=request.client.host if request.client else None
    )
    
    return {"message": "Parcel removed from trip"}


# ============ TRIP DOCUMENTS ROUTES ============

@router.get("/trips/{trip_id}/documents")
async def list_trip_documents(
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """List all documents for a trip"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    docs = await db.trip_documents.find(
        {"trip_id": trip_id},
        {"_id": 0}
    ).sort("uploaded_at", -1).to_list(100)
    
    # Enrich with uploader names
    result = []
    for doc in docs:
        uploader = await db.users.find_one({"id": doc.get("uploaded_by")}, {"name": 1, "_id": 0})
        result.append({
            **doc,
            "uploader_name": uploader.get("name") if uploader else "Unknown"
        })
    
    return result

@router.post("/trips/{trip_id}/documents")
async def upload_trip_document(
    trip_id: str,
    document: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Upload a document to a trip"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    doc = {
        "id": str(uuid.uuid4()),
        "trip_id": trip_id,
        "file_name": document.get("file_name"),
        "file_type": document.get("file_type"),
        "file_data": document.get("file_data"),
        "category": document.get("category", "Other"),
        "uploaded_by": user["id"],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.trip_documents.insert_one(doc)
    
    return {"id": doc["id"], "message": "Document uploaded successfully"}

@router.delete("/trips/{trip_id}/documents/{doc_id}")
async def delete_trip_document(
    trip_id: str,
    doc_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Delete a trip document"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    result = await db.trip_documents.delete_one({"id": doc_id, "trip_id": trip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {"message": "Document deleted"}

@router.get("/trips/{trip_id}/documents/{doc_id}/download")
async def download_trip_document(
    trip_id: str,
    doc_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get document data for download"""
    doc = await db.trip_documents.find_one({"id": doc_id, "trip_id": trip_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "file_name": doc["file_name"],
        "file_type": doc["file_type"],
        "file_data": doc["file_data"]
    }

# ============ TRIP DUPLICATE ROUTES ============

@router.post("/trips/{trip_id}/duplicate")
async def duplicate_trip(
    trip_id: str,
    request: Request,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Duplicate a trip (creates new trip with same route and settings, but no parcels)"""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Get next trip number
    trips = await db.trips.find(
        {"tenant_id": tenant_id, "trip_number": {"$regex": "^S\\d{1,4}$"}},
        {"trip_number": 1, "_id": 0}
    ).to_list(1000)
    
    max_num = 0
    for t in trips:
        try:
            num = int(t["trip_number"][1:])
            if num > max_num:
                max_num = num
        except (ValueError, IndexError, TypeError):
            continue
    
    new_trip_number = f"S{max_num + 1}"
    
    new_trip = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "trip_number": new_trip_number,
        "status": "planning",
        "route": trip.get("route", []),
        "notes": f"Duplicated from {trip.get('trip_number')}",
        "vehicle_id": trip.get("vehicle_id"),
        "driver_id": trip.get("driver_id"),
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.trips.insert_one(new_trip)
    
    return {"id": new_trip["id"], "trip_number": new_trip_number, "message": "Trip duplicated successfully"}



# ============ TRIP COMPLETION (SESSION P PART 1) ============

@router.post("/trips/{trip_id}/complete")
async def complete_trip(
    trip_id: str,
    request: Request,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """
    Mark trip as completed. Automatically:
    - Updates trip status to 'completed'
    - Updates all parcels with status 'loaded' or 'in_transit' to 'arrived'
    - Associates parcels with trip's destination warehouse
    - Records completion timestamp and user
    """
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Trip is already completed")
    
    # Get destination warehouse from trip route
    destination_warehouse_id = trip.get("destination_warehouse_id")
    
    # Update trip status to completed
    now = datetime.now(timezone.utc).isoformat()
    await db.trips.update_one(
        {"id": trip_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "completed",
            "completed_at": now,
            "completed_by": user["id"]
        }}
    )
    
    # Update all parcels from 'loaded' or 'in_transit' to 'arrived'
    parcel_update = {
        "status": "arrived",
        "arrived_at": now
    }
    
    # If trip has destination warehouse, associate parcels with it
    if destination_warehouse_id:
        parcel_update["warehouse_id"] = destination_warehouse_id
    
    result = await db.shipments.update_many(
        {
            "trip_id": trip_id,
            "tenant_id": tenant_id,
            "status": {"$in": ["loaded", "in_transit"]}
        },
        {"$set": parcel_update}
    )
    
    # Create audit log for trip completion
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.status_change,
        table_name="trips",
        record_id=trip_id,
        old_value={"status": trip.get("status")},
        new_value={"status": "completed"},
        ip_address=request.client.host if request.client else None
    )
    
    return {
        "message": f"Trip {trip.get('trip_number')} completed successfully",
        "parcels_updated": result.modified_count,
        "completed_at": now
    }


# ============ PACKING LIST ============

@router.get("/trips/{trip_id}/packing-list")
async def get_packing_list(trip_id: str, tenant_id: str = Depends(get_tenant_id)):
    """Get packing list items for a trip (from shipments)."""
    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")

    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0, "id": 1, "description": 1, "quantity": 1, "total_weight": 1}
    ).to_list(2000)

    items = sorted([{
        "shipment_id": s.get("id"),
        "description": s.get("description", "—"),
        "qty": s.get("quantity", 1) or 1,
        "kg": round(float(s.get("total_weight", 0) or 0), 2)
    } for s in shipments], key=lambda x: x["description"].lower())

    return {"trip_number": trip.get("trip_number"), "items": items}


@router.get("/trips/{trip_id}/packing-list/excel")
async def get_packing_list_excel(trip_id: str, tenant_id: str = Depends(get_tenant_id)):
    """Download packing list as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    trip_number = trip.get("trip_number", "TRIP")

    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0, "description": 1, "quantity": 1, "total_weight": 1}
    ).to_list(2000)
    items = sorted(shipments, key=lambda s: (s.get("description") or "").lower())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # Header style
    header_fill = PatternFill("solid", fgColor="1C2B1E")
    header_font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10

    for col, h in enumerate(["Description", "QTY", "KG"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_qty, total_kg = 0, 0.0
    for r, s in enumerate(items, 2):
        qty = int(s.get("quantity") or 1)
        kg = round(float(s.get("total_weight") or 0), 2)
        ws.cell(row=r, column=1, value=s.get("description", "—"))
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=kg)
        total_qty += qty
        total_kg += kg

    # Totals row
    tr = len(items) + 2
    bold = Font(bold=True)
    ws.cell(row=tr, column=1, value="TOTAL").font = bold
    ws.cell(row=tr, column=2, value=total_qty).font = bold
    ws.cell(row=tr, column=3, value=round(total_kg, 2)).font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Packing-List-{trip_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ DIGITAL MANIFEST ============

@router.get("/trips/{trip_id}/manifest/excel")
async def get_manifest_excel(trip_id: str, tenant_id: str = Depends(get_tenant_id)):
    """Download digital manifest as Excel with full parcel details."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    trip_number = trip.get("trip_number", "TRIP")

    # Fetch shipments sorted oldest to newest
    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(2000)

    # Batch fetch clients
    client_ids = list(set(s.get("client_id") for s in shipments if s.get("client_id")))
    clients = {}
    if client_ids:
        for c in await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0}).to_list(1000):
            clients[c["id"]] = c

    # Fetch line items for invoice amounts per shipment
    invoice_ids = list(set(s.get("invoice_id") for s in shipments if s.get("invoice_id")))
    invoice_line_items = {}
    if invoice_ids:
        for li in await db.invoice_line_items.find({"invoice_id": {"$in": invoice_ids}}, {"_id": 0}).to_list(5000):
            sid = li.get("shipment_id")
            if sid:
                invoice_line_items[sid] = li

    # KES rate from settings
    settings = await db.settings.find_one({"tenant_id": tenant_id})
    kes_rate = 6.67
    if settings and settings.get("currencies"):
        for cur in settings["currencies"]:
            if cur.get("code") == "KES":
                kes_rate = cur.get("exchange_rate", 6.67)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"

    headers = [
        "#", "Date(DD/MM)", "Sent By", "Sender Phone",
        "Primary Recipient", "Recip Phone", "Recip Email", "Recip Address",
        "Secondary Recipient", "Sec Address", "Description", "QTY", "KG",
        "L", "W", "H", "INV Number", "Destination", "Comments",
        "V(L*W*H/1000)", "CBM(L*W*H/1000000)", "Shipping Weight", "Item Price", "Item Price KSH",
        "Entry Time(DD/MM/YYYY)"
    ]

    # Row 1: trip title merged
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.value = f"{trip_number} WORKBOOK"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Row 2: column headers
    dark_fill = PatternFill("solid", fgColor="1C2B1E")
    white_bold = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = white_bold
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths
    col_widths = [5, 10, 20, 15, 20, 15, 20, 25, 20, 25, 30, 5, 8, 5, 5, 5, 15, 15, 20, 12, 15, 15, 12, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 30

    for row_num, s in enumerate(shipments, 1):
        client = clients.get(s.get("client_id", ""), {})
        li = invoice_line_items.get(s.get("id", ""), {})
        l = float(li.get("length_cm") or 0)
        w = float(li.get("width_cm") or 0)
        h = float(li.get("height_cm") or 0)
        vol_1000 = round(l * w * h / 1000, 4) if (l and w and h) else 0
        cbm = round(l * w * h / 1000000, 6) if (l and w and h) else 0
        weight = float(s.get("total_weight") or 0)
        ship_w = max(weight, l * w * h / 5000) if (l and w and h) else weight
        item_price = float(li.get("amount") or 0)
        item_price_kes = round(item_price * kes_rate, 2)
        created_at_str = s.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created_at_str.replace("Z", "+00:00")) if created_at_str else None
            date_ddmm = dt.strftime("%d/%m") if dt else ""
            entry_time = dt.strftime("%d/%m/%Y") if dt else ""
        except Exception:
            date_ddmm = ""
            entry_time = ""

        row_data = [
            row_num, date_ddmm,
            client.get("name", ""),
            client.get("phone", "") or client.get("whatsapp", ""),
            s.get("recipient_name", "") or s.get("recipient", ""),
            s.get("recipient_phone", ""),
            s.get("recipient_email", ""),
            s.get("recipient_address", "") or s.get("destination", ""),
            s.get("secondary_recipient", ""),
            s.get("secondary_address", ""),
            s.get("description", ""),
            int(s.get("quantity") or 1),
            round(weight, 2),
            l or "", w or "", h or "",
            s.get("invoice_number", ""),
            s.get("destination", ""),
            s.get("notes", ""),
            vol_1000 or "",
            cbm or "",
            f"{round(ship_w, 2):.2f} Kg",
            f"R {item_price:.2f}" if item_price else "",
            item_price_kes if item_price_kes else "",
            entry_time
        ]
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_num + 2, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Manifest-{trip_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ LABELS PDF ============

@router.get("/trips/{trip_id}/labels/pdf")
async def get_trip_labels_pdf(
    trip_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Generate labels PDF for all shipments in trip."""
    from services.pdf_service import generate_labels_pdf

    shipments = await db.shipments.find(
        {"trip_id": trip_id, "tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(None)

    if not shipments:
        raise HTTPException(404, "No shipments in this trip")

    shipment_ids = [s["id"] for s in shipments]
    pdf_buffer = await generate_labels_pdf(shipment_ids, tenant_id)

    trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id})
    trip_number = trip.get("trip_number", trip_id) if trip else trip_id

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=labels_{trip_number}.pdf"}
    )
