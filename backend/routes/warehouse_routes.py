"""
Warehouse routes for Servex Holdings backend.
Handles warehouse operations including parcel management and photo uploads.
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import base64

from database import db
from dependencies import get_current_user, get_tenant_id, build_warehouse_filter, check_warehouse_access
from models.enums import ShipmentStatus, AuditAction
from models.schemas import create_audit_log
from services.barcode_service import generate_barcode

router = APIRouter()

@router.get("/warehouses")
async def list_warehouses(
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """List all warehouses for tenant (filtered by user's allowed warehouses)"""
    # Get user's allowed warehouses
    allowed = user.get("allowed_warehouses")
    
    query = {"tenant_id": tenant_id}
    
    # If user has warehouse restrictions, only show allowed warehouses
    if allowed and len(allowed) > 0 and user.get("role") not in ["owner", "manager"]:
        query["id"] = {"$in": allowed}
    
    warehouses = await db.warehouses.find(query, {"_id": 0}).to_list(1000)
    return warehouses

@router.get("/warehouse/parcels")
async def list_warehouse_parcels(
    status: Optional[str] = None,
    client_id: Optional[str] = None,
    destination: Optional[str] = None,
    trip_id: Optional[str] = None,
    warehouse_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    weight_min: Optional[float] = None,
    weight_max: Optional[float] = None,
    search: Optional[str] = None,
    not_invoiced: Optional[bool] = None,
    sort_by: Optional[str] = "created_at",
    sort_order: Optional[str] = "desc",
    page: int = 1,
    page_size: int = 25,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """List parcels for warehouse manager with filtering, sorting and pagination.
    
    SECURITY: Applies warehouse-based filtering for restricted users.
    Users with allowed_warehouses set will only see parcels from those warehouses.
    """
    query = {"tenant_id": tenant_id}
    
    # SECURITY: Apply warehouse-based access control
    warehouse_filter = build_warehouse_filter(user)
    if warehouse_filter:
        # User has warehouse restrictions
        if warehouse_id:
            # Check if requested warehouse is in allowed list
            await check_warehouse_access(user, warehouse_id)
            query["warehouse_id"] = warehouse_id
        else:
            # Apply the warehouse filter
            query.update(warehouse_filter)
    elif warehouse_id:
        # No restrictions, use requested warehouse
        query["warehouse_id"] = warehouse_id
    
    # Apply filters
    if status and status != "all":
        if "," in status:
            query["status"] = {"$in": status.split(",")}
        else:
            query["status"] = status
    if client_id:
        query["client_id"] = client_id
    if destination:
        query["destination"] = destination
    if trip_id:
        if trip_id == "unassigned":
            query["trip_id"] = None
        else:
            query["trip_id"] = trip_id
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to + "T23:59:59"
        else:
            query["created_at"] = {"$lte": date_to + "T23:59:59"}
    if weight_min is not None:
        query["total_weight"] = {"$gte": weight_min}
    if weight_max is not None:
        if "total_weight" in query:
            query["total_weight"]["$lte"] = weight_max
        else:
            query["total_weight"] = {"$lte": weight_max}
    if search:
        # Search by description, destination, id, or client name
        # First, find client IDs that match the search term
        matching_clients = await db.clients.find(
            {"tenant_id": tenant_id, "name": {"$regex": search, "$options": "i"}},
            {"id": 1, "_id": 0}
        ).to_list(100)
        matching_client_ids = [c["id"] for c in matching_clients]
        
        search_conditions = [
            {"description": {"$regex": search, "$options": "i"}},
            {"destination": {"$regex": search, "$options": "i"}},
            {"id": {"$regex": search, "$options": "i"}}
        ]
        if matching_client_ids:
            search_conditions.append({"client_id": {"$in": matching_client_ids}})
        
        query["$or"] = search_conditions
    if not_invoiced:
        query["invoice_id"] = None
    
    # Count total
    total_count = await db.shipments.count_documents(query)
    
    # Sort
    sort_field = sort_by if sort_by in ["created_at", "total_weight", "destination", "status"] else "created_at"
    sort_direction = -1 if sort_order == "desc" else 1
    
    # Fetch shipments with pagination
    skip = (page - 1) * page_size
    shipments = await db.shipments.find(query, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with client, trip, user, and invoice data
    client_ids = list(set(s.get("client_id") for s in shipments if s.get("client_id")))
    trip_ids = list(set(s.get("trip_id") for s in shipments if s.get("trip_id")))
    user_ids = list(set(s.get("created_by") for s in shipments if s.get("created_by")))
    invoice_ids = list(set(s.get("invoice_id") for s in shipments if s.get("invoice_id")))
    
    clients = {}
    if client_ids:
        client_docs = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        clients = {c["id"]: c["name"] for c in client_docs}
    
    trips = {}
    if trip_ids:
        trip_docs = await db.trips.find({"id": {"$in": trip_ids}}, {"_id": 0, "id": 1, "trip_number": 1, "status": 1}).to_list(100)
        trips = {t["id"]: {"trip_number": t["trip_number"], "status": t["status"]} for t in trip_docs}
    
    users = {}
    if user_ids:
        user_docs = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        users = {u["id"]: u["name"] for u in user_docs}
    
    invoices = {}
    if invoice_ids:
        invoice_docs = await db.invoices.find(
            {"id": {"$in": invoice_ids}}, 
            {"_id": 0, "id": 1, "invoice_number": 1, "status": 1}
        ).to_list(100)
        invoices = {i["id"]: {"invoice_number": i["invoice_number"], "status": i["status"]} for i in invoice_docs}
    
    # Enrich shipments
    enriched = []
    for s in shipments:
        invoice_data = invoices.get(s.get("invoice_id"), {}) if s.get("invoice_id") else {}
        enriched.append({
            **s,
            "client_name": clients.get(s.get("client_id"), "Unknown"),
            "trip_number": trips.get(s.get("trip_id"), {}).get("trip_number") if s.get("trip_id") else None,
            "trip_status": trips.get(s.get("trip_id"), {}).get("status") if s.get("trip_id") else None,
            "staff_name": users.get(s.get("created_by"), "Unknown"),
            "invoice_number": invoice_data.get("invoice_number"),
            "invoice_status": invoice_data.get("status")
        })
    
    return {
        "items": enriched,
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": (total_count + page_size - 1) // page_size
    }

@router.get("/warehouse/parcels/{parcel_id}")
async def get_warehouse_parcel_detail(
    parcel_id: str,
    tenant_id: str = Depends(get_tenant_id)
):
    """Get detailed parcel info for warehouse detail modal"""
    shipment = await db.shipments.find_one(
        {"id": parcel_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    if not shipment:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    # Get pieces
    pieces = await db.shipment_pieces.find({"shipment_id": parcel_id}, {"_id": 0}).to_list(100)
    
    # Get client
    client = await db.clients.find_one({"id": shipment.get("client_id")}, {"_id": 0})
    
    # Get trip if assigned
    trip = None
    if shipment.get("trip_id"):
        trip = await db.trips.find_one({"id": shipment["trip_id"]}, {"_id": 0})
    
    # Get staff who created it
    staff = await db.users.find_one({"id": shipment.get("created_by")}, {"_id": 0, "id": 1, "name": 1})
    
    return {
        **shipment,
        "pieces": pieces,
        "client": client,
        "trip": trip,
        "staff": staff
    }

@router.put("/warehouse/parcels/bulk-status")
async def bulk_update_parcel_status(
    request: Request,
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Bulk update status for multiple parcels.
    
    When status is 'arrived', parcels are moved to the trip's destination_warehouse_id
    if one is configured. If no destination warehouse, warehouse_id is cleared (visible in "All Warehouses").
    
    When status is 'collected', parcels are removed from all warehouses (warehouse_id = None).
    """
    parcel_ids = data.get("parcel_ids", [])
    new_status = data.get("status")
    
    if not parcel_ids or not new_status:
        raise HTTPException(status_code=400, detail="parcel_ids and status required")
    
    update_data = {"status": new_status}
    
    # Handle special status transitions
    if new_status == "arrived":
        # When parcels arrive, check if trip has a destination warehouse
        # Get the first parcel to find the trip
        first_parcel = await db.shipments.find_one(
            {"id": parcel_ids[0], "tenant_id": tenant_id},
            {"_id": 0, "trip_id": 1}
        )
        
        if first_parcel and first_parcel.get("trip_id"):
            trip = await db.trips.find_one(
                {"id": first_parcel["trip_id"], "tenant_id": tenant_id},
                {"_id": 0, "destination_warehouse_id": 1}
            )
            
            if trip and trip.get("destination_warehouse_id"):
                # Move parcels to destination warehouse
                update_data["warehouse_id"] = trip["destination_warehouse_id"]
            else:
                # No destination warehouse - clear warehouse_id (visible in "All Warehouses")
                update_data["warehouse_id"] = None
    
    elif new_status == "collected":
        # Collected parcels are removed from all warehouses
        update_data["warehouse_id"] = None
        update_data["collected"] = True
        update_data["collected_by"] = user["id"]
        update_data["collected_at"] = datetime.now(timezone.utc).isoformat()
    
    # Update all parcels
    result = await db.shipments.update_many(
        {"id": {"$in": parcel_ids}, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Create audit logs for each
    for pid in parcel_ids:
        await create_audit_log(
            tenant_id=tenant_id,
            user_id=user["id"],
            action=AuditAction.status_change,
            table_name="shipments",
            record_id=pid,
            new_value={"status": new_status},
            ip_address=request.client.host if request.client else None
        )
    
    return {"message": f"Updated {result.modified_count} parcels", "count": result.modified_count}


# ============ COLLECTION WORKFLOW (SESSION G) ============

@router.get("/warehouse/parcels/{parcel_id}/collection-check")
async def check_collection_eligibility(
    parcel_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Check if a parcel can be collected. Returns warning if unpaid. (Session G P-16)"""
    parcel = await db.shipments.find_one({"id": parcel_id, "tenant_id": tenant_id}, {"_id": 0})
    if not parcel:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    # Check parcel status
    if parcel.get("status") != "arrived":
        return {
            "can_collect": False,
            "reason": "not_arrived",
            "message": f"Parcel has not arrived yet. Current status: {parcel.get('status', 'unknown')}"
        }
    
    # Check payment status
    invoice_id = parcel.get("invoice_id")
    if not invoice_id:
        return {
            "can_collect": True,
            "warning": "not_invoiced",
            "message": "This parcel has not been invoiced yet.",
            "requires_confirmation": True
        }
    
    invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": tenant_id}, {"_id": 0})
    if not invoice:
        return {
            "can_collect": True,
            "warning": "invoice_not_found",
            "message": "Invoice not found.",
            "requires_confirmation": True
        }
    
    invoice_status = invoice.get("status", "draft")
    total_amount = invoice.get("total", 0)
    paid_amount = invoice.get("paid_amount", 0)
    outstanding = total_amount - paid_amount
    
    if invoice_status == "paid":
        return {
            "can_collect": True,
            "payment_status": "paid",
            "message": "Invoice fully paid. Safe to collect."
        }
    elif invoice_status == "partial":
        return {
            "can_collect": True,
            "warning": "partial_payment",
            "payment_status": "partial",
            "total_amount": total_amount,
            "paid_amount": paid_amount,
            "outstanding": outstanding,
            "message": f"Partial payment: R {paid_amount:.2f} paid, R {outstanding:.2f} outstanding.",
            "requires_confirmation": True
        }
    else:
        return {
            "can_collect": True,
            "warning": "unpaid",
            "payment_status": invoice_status,
            "total_amount": total_amount,
            "outstanding": outstanding,
            "message": f"UNPAID: R {total_amount:.2f} outstanding. Collection requires manager approval.",
            "requires_confirmation": True,
            "requires_admin_notification": True
        }


@router.post("/warehouse/parcels/{parcel_id}/collect")
async def collect_parcel(
    parcel_id: str,
    request: Request,
    data: dict = None,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Marks a parcel as collected. Sends admin notification if unpaid. (Session G P-16)"""
    if data is None:
        data = {}
    
    confirmation_note = data.get("confirmation_note", "")
    
    parcel = await db.shipments.find_one({"id": parcel_id, "tenant_id": tenant_id}, {"_id": 0})
    if not parcel:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    if parcel.get("status") != "arrived":
        raise HTTPException(status_code=400, detail=f"Parcel status is '{parcel.get('status')}'. Only 'arrived' parcels can be collected.")
    
    # Check payment status for notification
    send_admin_notification = False
    notification_message = ""
    invoice_id = parcel.get("invoice_id")
    
    if invoice_id:
        invoice = await db.invoices.find_one({"id": invoice_id, "tenant_id": tenant_id}, {"_id": 0})
        if invoice and invoice.get("status") in ["draft", "sent", "overdue", "partial"]:
            send_admin_notification = True
            outstanding = invoice.get("total", 0) - invoice.get("paid_amount", 0)
            notification_message = (
                f"User {user.get('name', 'Unknown')} collected parcel {parcel_id[:8].upper()} "
                f"with R {outstanding:.2f} outstanding. "
                f"Invoice: {invoice.get('invoice_number', 'N/A')}. "
                f"Note: {confirmation_note or 'None'}"
            )
    
    # Update parcel status
    now = datetime.now(timezone.utc).isoformat()
    await db.shipments.update_one(
        {"id": parcel_id, "tenant_id": tenant_id},
        {"$set": {
            "status": "collected",
            "collected_at": now,
            "collected_by": user["id"],
            "collection_note": confirmation_note,
            "locked": True,  # Lock parcel to prevent further editing
            "updated_at": now
        }}
    )
    
    # Create audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.status_change,
        table_name="shipments",
        record_id=parcel_id,
        old_value={"status": "arrived"},
        new_value={"status": "collected"},
        ip_address=request.client.host if request.client else None
    )
    
    # Send admin notification if unpaid
    admin_notified = False
    if send_admin_notification:
        admins = await db.users.find(
            {"tenant_id": tenant_id, "role": {"$in": ["owner", "manager"]}},
            {"_id": 0}
        ).to_list(50)
        
        for admin in admins:
            await db.notifications.insert_one({
                "id": str(__import__('uuid').uuid4()),
                "tenant_id": tenant_id,
                "user_id": admin["id"],
                "type": "collection_warning",
                "title": "Unpaid Parcel Collected",
                "message": notification_message,
                "parcel_id": parcel_id,
                "invoice_id": invoice_id,
                "read": False,
                "created_at": now
            })
        admin_notified = True
    
    return {
        "success": True,
        "parcel_id": parcel_id,
        "collected_at": now,
        "admin_notified": admin_notified
    }


@router.post("/warehouse/scan-collect")
async def scan_and_collect_parcel(
    request: Request,
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Scan a barcode or parcel ID and mark it as collected.
    
    Supports multiple lookup methods:
    1. Full barcode (e.g., S123-001-01)
    2. Partial shipment ID (e.g., E1DF9124 - first 8 chars)
    3. Full shipment ID (UUID format)
    
    Only parcels with 'arrived' status can be collected.
    """
    barcode = data.get("barcode", "").strip()
    
    if not barcode:
        raise HTTPException(status_code=400, detail="Barcode or parcel ID required")
    
    # Find the shipment
    shipment = None
    barcode_upper = barcode.upper()
    barcode_lower = barcode.lower()
    
    # First try piece barcode
    piece = await db.shipment_pieces.find_one({"barcode": barcode}, {"_id": 0})
    if piece:
        shipment = await db.shipments.find_one(
            {"id": piece["shipment_id"], "tenant_id": tenant_id},
            {"_id": 0}
        )
    
    if not shipment:
        # Try full ID match
        shipment = await db.shipments.find_one(
            {"id": barcode_lower, "tenant_id": tenant_id},
            {"_id": 0}
        )
    
    if not shipment:
        # Try partial ID match (first 8 characters)
        shipments = await db.shipments.find(
            {"tenant_id": tenant_id},
            {"_id": 0, "id": 1}
        ).to_list(10000)
        
        for s in shipments:
            if s["id"][:8].upper() == barcode_upper or s["id"].upper().startswith(barcode_upper):
                shipment = await db.shipments.find_one(
                    {"id": s["id"], "tenant_id": tenant_id},
                    {"_id": 0}
                )
                break
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    # Check if parcel can be collected (must be 'arrived' status)
    if shipment.get("status") != "arrived":
        current_status = shipment.get("status", "unknown")
        raise HTTPException(
            status_code=400, 
            detail=f"Parcel status is '{current_status}'. Only parcels with 'arrived' status can be collected."
        )
    
    # Mark as collected
    now = datetime.now(timezone.utc).isoformat()
    update_data = {
        "status": "collected",
        "collected": True,
        "collected_by": user["id"],
        "collected_at": now,
        "locked": True,  # Lock parcel to prevent further editing
        "warehouse_id": None  # Remove from all warehouses
    }
    
    await db.shipments.update_one(
        {"id": shipment["id"], "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Create audit log
    await create_audit_log(
        tenant_id=tenant_id,
        user_id=user["id"],
        action=AuditAction.status_change,
        table_name="shipments",
        record_id=shipment["id"],
        old_value={"status": shipment.get("status")},
        new_value={"status": "collected"},
        ip_address=request.client.host if request.client else None
    )
    
    # Get client info for response
    client = await db.clients.find_one({"id": shipment.get("client_id")}, {"_id": 0, "name": 1})
    
    return {
        "message": "Parcel marked as collected",
        "parcel_id": shipment["id"],
        "description": shipment.get("description"),
        "client_name": client.get("name") if client else "Unknown",
        "collected_at": now
    }

@router.put("/warehouse/parcels/bulk-assign-trip")
async def bulk_assign_parcels_to_trip(
    request: Request,
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Bulk assign multiple parcels to a trip"""
    parcel_ids = data.get("parcel_ids", [])
    trip_id = data.get("trip_id")
    
    if not parcel_ids:
        raise HTTPException(status_code=400, detail="parcel_ids required")
    
    # Verify trip exists if assigning
    if trip_id:
        trip = await db.trips.find_one({"id": trip_id, "tenant_id": tenant_id}, {"_id": 0})
        if not trip:
            raise HTTPException(status_code=404, detail="Trip not found")
        if trip.get("locked_at"):
            raise HTTPException(status_code=403, detail="Cannot assign to locked trip")
    
    # Update all parcels - also set destination from trip route
    destination = None
    if trip_id and trip.get("route") and len(trip["route"]) > 0:
        destination = trip["route"][-1]

    update_data = {"trip_id": trip_id, "status": "staged" if trip_id else "warehouse"}
    if destination:
        update_data["destination"] = destination
    elif not trip_id:
        update_data["destination"] = "No Trip"
    result = await db.shipments.update_many(
        {"id": {"$in": parcel_ids}, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Update barcodes if assigning to a trip
    if trip_id:
        for parcel_id in parcel_ids:
            pieces = await db.shipment_pieces.find({"shipment_id": parcel_id}, {"_id": 0}).to_list(100)
            shipment_count = await db.shipments.count_documents({
                "tenant_id": tenant_id,
                "trip_id": trip_id
            })
            for piece in pieces:
                new_barcode = generate_barcode(trip["trip_number"], shipment_count, piece["piece_number"])
                await db.shipment_pieces.update_one(
                    {"id": piece["id"]},
                    {"$set": {"barcode": new_barcode}}
                )
    
    return {"message": f"Assigned {result.modified_count} parcels to trip", "count": result.modified_count}

@router.delete("/warehouse/parcels/bulk-delete")
async def bulk_delete_parcels(
    request: Request,
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Bulk delete multiple parcels"""
    parcel_ids = data.get("parcel_ids", [])
    
    if not parcel_ids:
        raise HTTPException(status_code=400, detail="parcel_ids required")
    
    # Delete pieces first
    for parcel_id in parcel_ids:
        await db.shipment_pieces.delete_many({"shipment_id": parcel_id})
    
    # Delete parcels
    result = await db.shipments.delete_many(
        {"id": {"$in": parcel_ids}, "tenant_id": tenant_id}
    )
    
    # Audit logs
    for pid in parcel_ids:
        await create_audit_log(
            tenant_id=tenant_id,
            user_id=user["id"],
            action=AuditAction.delete,
            table_name="shipments",
            record_id=pid,
            ip_address=request.client.host if request.client else None
        )
    
    return {"message": f"Deleted {result.deleted_count} parcels", "count": result.deleted_count}

@router.put("/warehouse/parcels/bulk-collect")
async def bulk_collect_parcels(
    request: Request,
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Mark multiple parcels as collected (for arrivals that have been picked up)"""
    parcel_ids = data.get("parcel_ids", [])
    
    if not parcel_ids:
        raise HTTPException(status_code=400, detail="parcel_ids required")
    
    # Only allow collection of parcels that have status "arrived"
    arrived_parcels = await db.shipments.find(
        {"id": {"$in": parcel_ids}, "tenant_id": tenant_id, "status": "arrived"},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    
    valid_ids = [p["id"] for p in arrived_parcels]
    
    if not valid_ids:
        raise HTTPException(status_code=400, detail="No parcels with 'arrived' status to collect")
    
    # Update parcels to collected status
    now = datetime.now(timezone.utc).isoformat()
    update_data = {
        "status": "collected",
        "collected": True,
        "collected_by": user["id"],
        "collected_at": now,
        "locked": True  # Lock parcels to prevent further editing
    }
    
    result = await db.shipments.update_many(
        {"id": {"$in": valid_ids}, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    # Audit logs
    for pid in valid_ids:
        await create_audit_log(
            tenant_id=tenant_id,
            user_id=user["id"],
            action=AuditAction.status_change,
            table_name="shipments",
            record_id=pid,
            new_value={"status": "collected"},
            ip_address=request.client.host if request.client else None
        )
    
    skipped = len(parcel_ids) - len(valid_ids)
    message = f"Marked {result.modified_count} parcel(s) as collected"
    if skipped > 0:
        message += f" ({skipped} skipped - only 'arrived' parcels can be collected)"
    
    return {"message": message, "count": result.modified_count}

@router.get("/warehouse/filters")
async def get_warehouse_filter_options(tenant_id: str = Depends(get_tenant_id)):
    """Get available filter options for warehouse manager"""
    # Get unique destinations
    destinations = await db.shipments.distinct("destination", {"tenant_id": tenant_id})
    
    # Get clients with shipments
    client_ids = await db.shipments.distinct("client_id", {"tenant_id": tenant_id})
    clients = await db.clients.find(
        {"id": {"$in": client_ids}, "tenant_id": tenant_id},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    
    # Get active trips
    trips = await db.trips.find(
        {"tenant_id": tenant_id, "status": {"$nin": ["closed", "delivered"]}},
        {"_id": 0, "id": 1, "trip_number": 1, "status": 1}
    ).to_list(100)
    
    return {
        "destinations": destinations,
        "clients": clients,
        "trips": trips,
        "statuses": ["warehouse", "staged", "loaded", "in_transit", "delivered"]
    }

@router.post("/warehouse/parcels/{parcel_id}/photos")
async def upload_parcel_photo(
    parcel_id: str,
    piece_id: Optional[str] = None,
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Upload a photo to a parcel piece"""
    # Verify parcel exists
    shipment = await db.shipments.find_one({"id": parcel_id, "tenant_id": tenant_id})
    if not shipment:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    # Get pieces for this parcel
    pieces = await db.shipment_pieces.find({"shipment_id": parcel_id}, {"_id": 0}).to_list(100)
    if not pieces:
        raise HTTPException(status_code=404, detail="No pieces found for this parcel")
    
    # Determine which piece to update
    target_piece = None
    if piece_id:
        target_piece = next((p for p in pieces if p["id"] == piece_id), None)
        if not target_piece:
            raise HTTPException(status_code=404, detail="Piece not found")
    else:
        # Find first piece without photo or first piece
        target_piece = next((p for p in pieces if not p.get("photo_url")), pieces[0])
    
    # Read file content and convert to base64 data URL
    content = await file.read()
    content_type = file.content_type or "image/jpeg"
    
    base64_data = base64.b64encode(content).decode('utf-8')
    photo_url = f"data:{content_type};base64,{base64_data}"
    
    # Update piece with photo
    await db.shipment_pieces.update_one(
        {"id": target_piece["id"]},
        {"$set": {"photo_url": photo_url}}
    )
    
    return {
        "message": "Photo uploaded successfully",
        "piece_id": target_piece["id"],
        "photo_url": photo_url
    }

@router.delete("/warehouse/parcels/{parcel_id}/photos/{piece_id}")
async def delete_parcel_photo(
    parcel_id: str,
    piece_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Delete a photo from a parcel piece"""
    # Verify parcel exists
    shipment = await db.shipments.find_one({"id": parcel_id, "tenant_id": tenant_id})
    if not shipment:
        raise HTTPException(status_code=404, detail="Parcel not found")
    
    # Verify piece exists
    piece = await db.shipment_pieces.find_one({"id": piece_id, "shipment_id": parcel_id})
    if not piece:
        raise HTTPException(status_code=404, detail="Piece not found")
    
    # Remove photo URL
    await db.shipment_pieces.update_one(
        {"id": piece_id},
        {"$set": {"photo_url": None}}
    )
    
    return {"message": "Photo deleted successfully", "piece_id": piece_id}

# ============ WAREHOUSE CRUD ============

from pydantic import BaseModel
import uuid

class WarehouseCreate(BaseModel):
    name: str
    location: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    status: str = "active"

class WarehouseUpdate(BaseModel):
    name: Optional[str] = None
    location: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    status: Optional[str] = None

@router.post("/warehouses")
async def create_warehouse(
    warehouse_data: WarehouseCreate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Create a new warehouse"""
    warehouse = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "name": warehouse_data.name,
        "location": warehouse_data.location,
        "contact_person": warehouse_data.contact_person,
        "phone": warehouse_data.phone,
        "status": warehouse_data.status,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    
    await db.warehouses.insert_one(warehouse)
    
    # Return without _id
    if "_id" in warehouse:
        del warehouse["_id"]
    return warehouse

@router.put("/warehouses/{warehouse_id}")
async def update_warehouse(
    warehouse_id: str,
    warehouse_data: WarehouseUpdate,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Update a warehouse"""
    existing = await db.warehouses.find_one(
        {"id": warehouse_id, "tenant_id": tenant_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    update_dict = {k: v for k, v in warehouse_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.warehouses.update_one(
            {"id": warehouse_id, "tenant_id": tenant_id},
            {"$set": update_dict}
        )
    
    warehouse = await db.warehouses.find_one(
        {"id": warehouse_id, "tenant_id": tenant_id},
        {"_id": 0}
    )
    return warehouse

@router.delete("/warehouses/{warehouse_id}")
async def delete_warehouse(
    warehouse_id: str,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Delete a warehouse"""
    existing = await db.warehouses.find_one(
        {"id": warehouse_id, "tenant_id": tenant_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Warehouse not found")
    
    # Check if warehouse has parcels
    parcel_count = await db.shipments.count_documents(
        {"warehouse_id": warehouse_id, "tenant_id": tenant_id}
    )
    if parcel_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete warehouse with {parcel_count} parcels assigned"
        )
    
    await db.warehouses.delete_one({"id": warehouse_id, "tenant_id": tenant_id})
    return {"message": "Warehouse deleted successfully"}

@router.post("/warehouses/create-defaults")
async def create_default_warehouses(
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Create default warehouses (Johannesburg and Nairobi)"""
    defaults = [
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Johannesburg Warehouse",
            "location": "Johannesburg, South Africa",
            "contact_person": None,
            "phone": None,
            "status": "active",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"]
        },
        {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant_id,
            "name": "Nairobi Warehouse",
            "location": "Nairobi, Kenya",
            "contact_person": None,
            "phone": None,
            "status": "active",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"]
        }
    ]
    
    created = []
    for warehouse in defaults:
        # Check if warehouse with same name exists
        existing = await db.warehouses.find_one(
            {"tenant_id": tenant_id, "name": warehouse["name"]}
        )
        if not existing:
            await db.warehouses.insert_one(warehouse)
            created.append(warehouse["name"])
    
    if created:
        return {"message": f"Created warehouses: {', '.join(created)}", "created": created}
    else:
        return {"message": "Default warehouses already exist", "created": []}

# ============ HEALTH CHECK ============


# ============ WAREHOUSE LABELS PDF ============

@router.post("/warehouse/labels/pdf")
async def get_warehouse_labels_pdf(
    data: dict,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Generate labels PDF for selected warehouse shipments."""
    from services.pdf_service import generate_labels_pdf

    shipment_ids = data.get("shipment_ids", [])
    if not shipment_ids:
        raise HTTPException(400, "No shipments selected")

    pdf_buffer = await generate_labels_pdf(shipment_ids, tenant_id)

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=warehouse_labels.pdf"}
    )


# ============ WAREHOUSE EXCEL EXPORT ============

@router.get("/warehouse/export/excel")
async def export_warehouse_excel(
    warehouse_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id),
    user: dict = Depends(get_current_user)
):
    """Export warehouse parcels as Excel in Digital Manifest format (24 columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    query = {"tenant_id": tenant_id}

    warehouse_filter = build_warehouse_filter(user)
    if warehouse_filter:
        if warehouse_id and warehouse_id != "all":
            await check_warehouse_access(user, warehouse_id)
            query["warehouse_id"] = warehouse_id
        else:
            query.update(warehouse_filter)
    elif warehouse_id and warehouse_id != "all":
        query["warehouse_id"] = warehouse_id

    if status and status != "all":
        if "," in status:
            query["status"] = {"$in": status.split(",")}
        else:
            query["status"] = status

    if search:
        query["$or"] = [
            {"description": {"$regex": search, "$options": "i"}},
            {"recipient": {"$regex": search, "$options": "i"}},
            {"id": {"$regex": search, "$options": "i"}}
        ]

    parcels = await db.shipments.find(query, {"_id": 0}).sort("created_at", 1).to_list(None)

    if not parcels:
        # Return empty Excel instead of 404
        from openpyxl import Workbook as WB2
        wb_empty = WB2()
        ws_empty = wb_empty.active
        ws_empty.title = "Warehouse Export"
        ws_empty['A1'] = "No parcels found matching filters"
        empty_output = BytesIO()
        wb_empty.save(empty_output)
        empty_output.seek(0)
        return StreamingResponse(
            empty_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=warehouse_export_empty.xlsx"}
        )

    client_ids = list(set(p.get("client_id") for p in parcels if p.get("client_id")))
    clients = await db.clients.find({"id": {"$in": client_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
    clients_map = {c["id"]: c["name"] for c in clients}

    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Export"

    ws['A1'] = "WAREHOUSE EXPORT - DIGITAL MANIFEST FORMAT"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Warehouse: {warehouse_id if warehouse_id and warehouse_id != 'all' else 'All'}"
    ws['A3'] = f"Status: {status if status and status != 'all' else 'All'}"
    ws['A4'] = f"Search: {search if search else 'None'}"
    ws['A5'] = f"Exported: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"

    headers = [
        "#", "Parcel ID", "Barcode", "Client", "Recipient", "Description",
        "Pieces", "L (cm)", "W (cm)", "H (cm)", "Vol Weight", "Actual Weight (kg)",
        "Shipping Weight", "CBM", "Destination", "Trip", "Invoice #",
        "Invoice Status", "Status", "Warehouse", "Date In", "Date Out",
        "Collected Date", "Notes"
    ]

    header_row = 7
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, parcel in enumerate(parcels, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - header_row)
        ws.cell(row=row_idx, column=2, value=parcel.get("id", ""))
        ws.cell(row=row_idx, column=3, value=parcel.get("barcode", ""))
        ws.cell(row=row_idx, column=4, value=clients_map.get(parcel.get("client_id"), "Unknown"))
        ws.cell(row=row_idx, column=5, value=parcel.get("recipient", ""))
        ws.cell(row=row_idx, column=6, value=parcel.get("description", ""))
        ws.cell(row=row_idx, column=7, value=parcel.get("total_pieces", 1))
        ws.cell(row=row_idx, column=8, value=parcel.get("length_cm", 0))
        ws.cell(row=row_idx, column=9, value=parcel.get("width_cm", 0))
        ws.cell(row=row_idx, column=10, value=parcel.get("height_cm", 0))
        vol_weight = 0
        l, w, h = parcel.get("length_cm", 0) or 0, parcel.get("width_cm", 0) or 0, parcel.get("height_cm", 0) or 0
        if l and w and h:
            vol_weight = round((l * w * h) / 5000, 2)
        ws.cell(row=row_idx, column=11, value=vol_weight)
        ws.cell(row=row_idx, column=12, value=round(parcel.get("total_weight", 0) or 0, 2))
        shipping_weight = max(parcel.get("total_weight", 0) or 0, vol_weight)
        ws.cell(row=row_idx, column=13, value=round(shipping_weight, 2))
        ws.cell(row=row_idx, column=14, value=round(parcel.get("total_cbm", 0) or 0, 4))
        ws.cell(row=row_idx, column=15, value=parcel.get("destination", ""))
        ws.cell(row=row_idx, column=16, value=parcel.get("trip_number", ""))
        ws.cell(row=row_idx, column=17, value=parcel.get("invoice_number", ""))
        ws.cell(row=row_idx, column=18, value=parcel.get("invoice_status", ""))
        ws.cell(row=row_idx, column=19, value=parcel.get("status", ""))
        ws.cell(row=row_idx, column=20, value=parcel.get("warehouse_name", ""))
        ws.cell(row=row_idx, column=21, value=str(parcel.get("created_at", ""))[:10])
        ws.cell(row=row_idx, column=22, value=str(parcel.get("loaded_at", ""))[:10] if parcel.get("loaded_at") else "")
        ws.cell(row=row_idx, column=23, value=str(parcel.get("collected_at", ""))[:10] if parcel.get("collected_at") else "")
        ws.cell(row=row_idx, column=24, value=parcel.get("notes", ""))

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"warehouse_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
